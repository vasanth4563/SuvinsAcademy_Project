import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

EXCEL_FILE = os.getenv("EXCEL_FILE", "enquiries_v2.xlsx")

HEADERS = [
    "Submitted Date & Time",
    "Full Name",
    "Mobile Number",
    "School Name",
    "Class",
    "Standard",
]

HEADER_FILL   = PatternFill("solid", fgColor="192B5C")   # Navy blue
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ROW_ALIGN     = Alignment(horizontal="left",   vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COL_WIDTHS = [22, 25, 18, 30, 8, 15]


def _apply_header(ws):
    """Write and style the header row."""
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 30


def _style_data_row(ws, row_idx: int):
    """Apply alternating row colour and border to a data row."""
    fill_color = "EEF2FA" if row_idx % 2 == 0 else "FFFFFF"
    fill = PatternFill("solid", fgColor=fill_color)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill      = fill
        cell.alignment = ROW_ALIGN
        cell.border    = THIN_BORDER
    ws.row_dimensions[row_idx].height = 20


def ensure_excel():
    """Create the Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Enquiries"
        _apply_header(ws)
        ws.freeze_panes = "A2"           # Freeze header row
        wb.save(EXCEL_FILE)
        print(f"[OK]  Created new Excel file: {EXCEL_FILE}")


def append_enquiry(
    full_name: str,
    mobile_number: str,
    school_name: str,
    class_name: str,
    standard: str,
    submitted_at: datetime,
):
    """Append a new enquiry row to the Excel file."""
    ensure_excel()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    next_row = ws.max_row + 1
    row_data = [
        submitted_at.strftime("%Y-%m-%d %H:%M:%S"),
        full_name,
        mobile_number,
        school_name,
        class_name,
        standard,
    ]

    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=next_row, column=col_idx, value=value)

    _style_data_row(ws, next_row)
    wb.save(EXCEL_FILE)
    print(f"[OK]  Row {next_row - 1} appended to Excel: {full_name}")
